#!/usr/bin/env python3
"""Number hygiene in the Excel dashboard.

Seen on a real corpus on 2026-06-11: a cell showing 70.90000000000001. The
workbook used to write view floats into cells verbatim, with the default
``General`` number format, so any value that escaped the view's rounding (or
any future field added unrounded) surfaced its binary float tail in the grid
and in the formula bar. The renderer now rounds every numeric cell at write
time by metric kind (counts 0 decimals, 0-100 scores 1, 0-1 ratios 2) and
sets the matching Excel number format, so a tail can never surface. These
tests feed the renderer hostile floats directly and inspect the stored cell
values and formats.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import pytest

PLUGIN_ROOT = Path(__file__).resolve().parent.parent
if str(PLUGIN_ROOT) not in sys.path:
    sys.path.insert(0, str(PLUGIN_ROOT))

from modules.output_forge import _parse_timestamp  # noqa: E402
from modules.output_forge.brand import resolve_brand  # noqa: E402

# A float repr an analyst should ever see in a cell: at most two decimals.
_CLEAN_FLOAT = re.compile(r"^-?\d+\.\d{1,2}$")


def _hostile_view() -> dict:
    """A view whose every numeric field carries a binary float tail."""
    return {
        "header": {"label": "Hygiene audit"},
        "corpus": {
            "total_keywords": 170,
            "aio_eligibility_share": 0.7090000000000001,
            "geo_opportunity_share": 0.30000000000000004,
            "demand_opportunity_score": 62.71570000000001,
            "intent_split": [{"intent": "informational", "count": 120}],
        },
        "clusters": [
            {
                "head": "running shoes",
                "size": 12,
                "volume_total": 12345.000000000002,
                "dominant_intent": "commercial",
                "topical_authority": 0.7000000000000001,
                "cover_at_tau": 0.6100000000000001,
                "expected_readiness": 70.90000000000001,
                "expected_share": 33.33333333333333,
                "winnable_band": [1200, 2400],
                "observed_current_clicks": 70.90000000000001,
                "observed_citation_share": 0.12000000000000001,
            }
        ],
        "winnable": {
            "by_intent": [
                {"intent": "commercial", "winnable_band": [10.000000000000002, 20]}
            ],
            "portfolio_winnable_band": [30, 60.00000000000001],
        },
        "gaps": [
            {
                "entity": "trail",
                "demand_volume": 9999.000000000001,
                "suggested_cluster_head": "running shoes",
            }
        ],
    }


def _render(tmp_path: Path):
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    from modules.output_forge.dashboard_xlsx import render_workbook

    out = tmp_path / "dash.xlsx"
    ts = _parse_timestamp("2026-06-08T00:00:00Z")
    render_workbook(_hostile_view(), resolve_brand(None), out, ts)
    return load_workbook(out)


def test_no_cell_shows_a_float_tail(tmp_path: Path) -> None:
    """The sweep: every numeric cell is clean and carries a real number format."""
    wb = _render(tmp_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                v = cell.value
                if isinstance(v, bool) or not isinstance(v, (int, float)):
                    continue
                if isinstance(v, float):
                    assert _CLEAN_FLOAT.match(repr(v)), (
                        f"{ws.title}!{cell.coordinate} stores {v!r}"
                    )
                assert cell.number_format != "General", (
                    f"{ws.title}!{cell.coordinate} has no number format"
                )


def test_metric_kinds_round_and_format_as_declared(tmp_path: Path) -> None:
    """Spot checks per metric kind: counts 0 decimals, scores 1, ratios 2."""
    wb = _render(tmp_path)
    clusters = wb["Clusters"]
    row = list(clusters.iter_rows(min_row=2, max_row=2))[0]
    by_header = {clusters.cell(row=1, column=c.column).value: c for c in row}

    # Counts: integer value, thousands format.
    for header, expected in (
        ("Volume", 12345),
        ("Winnable low", 1200),
        ("Observed clicks", 71),
    ):
        cell = by_header[header]
        assert cell.value == expected and isinstance(cell.value, int), (
            f"{header}: {cell.value!r}"
        )
        assert cell.number_format == "#,##0", f"{header}: {cell.number_format!r}"

    # 0-100 scores: one decimal.
    for header, expected in (
        ("Expected readiness", 70.9),
        ("Expected share", 33.3),
    ):
        cell = by_header[header]
        assert cell.value == expected, f"{header}: {cell.value!r}"
        assert cell.number_format == "0.0", f"{header}: {cell.number_format!r}"

    # 0-1 ratios: two decimals.
    for header, expected in (
        ("Topical authority", 0.7),
        ("Cover@tau", 0.61),
        ("Observed citation share", 0.12),
    ):
        cell = by_header[header]
        assert cell.value == expected, f"{header}: {cell.value!r}"
        assert cell.number_format == "0.00", f"{header}: {cell.number_format!r}"


def test_overview_values_round_by_row_kind(tmp_path: Path) -> None:
    """The Overview metric column mixes kinds, so the kind is set per row."""
    wb = _render(tmp_path)
    ws = wb["Overview"]
    values = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2)
        for r in range(1, ws.max_row + 1)
    }
    assert values["Keywords analysed"].value == 170
    assert values["AI Overview eligibility share"].value == 0.71
    assert values["AI Overview eligibility share"].number_format == "0.00"
    assert values["Generative-engine opportunity share"].value == 0.3
    assert values["Demand opportunity score"].value == 62.7
    assert values["Demand opportunity score"].number_format == "0.0"
