#!/usr/bin/env python3
"""Excel dashboard renderer (.xlsx) behind openpyxl.

Turns the report view into a workbook an analyst can sort and filter: an Overview
sheet with the headline figures, a Clusters sheet with every per-cluster metric, a
Gaps sheet, and a Winnable sheet split by intent. The document properties and dates
are pinned to the run timestamp so the file is byte-stable across runs once the port
normalises the archive.

This path needs ``openpyxl`` (an optional backend). The caller checks
``ooxml_capabilities()`` and skips this format cleanly when the package is absent.
The rendering is exercised in continuous integration; locally the suite degrades to
the HTML dashboard.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from .brand import office_hex


# Leading characters that make Excel treat a cell as a formula. ``=`` is rendered as
# a live formula by openpyxl itself; ``+ - @`` and a leading tab/CR/newline are read
# as formulas by Excel on entry. A cluster head, entity, or title taken from the input
# could smuggle one in, so the value is neutralised before it reaches a cell.
_DANGEROUS_LEADERS = ("=", "+", "-", "@", "\t", "\r", "\n")


def _safe_cell(value: Any) -> Any:
    """Neutralise spreadsheet formula injection in a string cell value (CWE-1236).

    Prefixes a single apostrophe to any string that starts with a formula leader, so
    Excel reads the cell as literal text instead of evaluating it. Real keyword data
    never starts with one of these characters, so legitimate output is untouched;
    only a hostile value gets the guard. Non-string values pass through unchanged, so
    the numeric metrics keep their type.
    """
    if isinstance(value, str) and value[:1] in _DANGEROUS_LEADERS:
        return "'" + value
    return value


# Decimals and Excel number format per metric kind, declared once. The view already
# rounds most figures, but the workbook is the last gate: every numeric cell is
# rounded again at write time AND carries the matching number format, so a binary
# float tail (70.90000000000001 was seen on a real corpus) can never surface, not in
# the grid and not in the formula bar. Counts (keywords, volumes, clicks, click
# bands, demand) show whole numbers; 0-100 scores show one decimal; 0-1 ratios show
# two. Rounding is round-half-even (Python's built-in ``round``), the same
# convention the view uses in ``model.py``.
_COUNT = ("#,##0", 0)
_SCORE = ("0.0", 1)
_RATIO = ("0.00", 2)


def _style_numbers(ws, kinds: list[tuple[str, int] | None]) -> None:
    """Round and format the numeric cells of the last appended row, by kind.

    ``kinds`` aligns with the row's columns; ``None`` leaves a column untouched
    (text columns). Non-numeric values (strings, ``None``) pass through.
    """
    for cell, kind in zip(ws[ws.max_row], kinds):
        if kind is None:
            continue
        value = cell.value
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            continue
        fmt, decimals = kind
        cell.value = round(float(value), decimals) if decimals else round(value)
        cell.number_format = fmt


def _header_row(ws, headers: list[str], brand: dict[str, Any]) -> None:
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill("solid", fgColor=office_hex(brand["colors"]["primary"]))
    font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.font = font


def render_workbook(
    view: dict[str, Any], brand: dict[str, Any], out_path: Path, timestamp: datetime
) -> None:
    """Render the Excel dashboard to ``out_path``. Requires openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    header = view["header"]
    title_text = header["label"] or f"{brand['name']} keyword and demand audit"
    wb = Workbook()

    # Overview sheet.
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = _safe_cell(title_text)
    ws["A1"].font = Font(
        bold=True, size=14, color=office_hex(brand["colors"]["primary"])
    )
    ws["A2"] = _safe_cell(f"{brand['name']} | {brand['tagline']}")
    corpus = view["corpus"]
    ws.append([])
    _header_row(ws, ["Metric", "Value"], brand)
    for label, value, kind in (
        ("Keywords analysed", corpus["total_keywords"], _COUNT),
        ("AI Overview eligibility share", corpus["aio_eligibility_share"], _RATIO),
        (
            "Generative-engine opportunity share",
            corpus["geo_opportunity_share"],
            _RATIO,
        ),
        ("Demand opportunity score", corpus["demand_opportunity_score"], _SCORE),
    ):
        ws.append([label, value])
        _style_numbers(ws, [None, kind])
    ws.append([])
    _header_row(ws, ["Search intent", "Keywords"], brand)
    for row in corpus["intent_split"]:
        ws.append([_safe_cell(row["intent"].replace("_", " ")), row["count"]])
        _style_numbers(ws, [None, _COUNT])

    # Clusters sheet.
    cs = wb.create_sheet("Clusters")
    cols = [
        "Cluster",
        "Size",
        "Volume",
        "Intent",
        "Topical authority",
        "Cover@tau",
        "Expected readiness",
        "Expected share",
        "Winnable low",
        "Winnable high",
    ]
    has_obs = any("observed_current_clicks" in c for c in view["clusters"])
    if has_obs:
        cols += ["Observed clicks", "Observed citation share"]
    _header_row(cs, cols, brand)
    cluster_kinds: list[tuple[str, int] | None] = [
        None,
        _COUNT,
        _COUNT,
        None,
        _RATIO,
        _RATIO,
        _SCORE,
        _SCORE,
        _COUNT,
        _COUNT,
    ]
    if has_obs:
        cluster_kinds += [_COUNT, _RATIO]
    for c in view["clusters"]:
        band = c.get("winnable_band") or [None, None]
        row = [
            _safe_cell(c["head"]),
            c.get("size"),
            c.get("volume_total"),
            _safe_cell(str(c.get("dominant_intent", "")).replace("_", " ")),
            c.get("topical_authority"),
            c.get("cover_at_tau"),
            c.get("expected_readiness"),
            c.get("expected_share"),
            band[0],
            band[1],
        ]
        if has_obs:
            row += [c.get("observed_current_clicks"), c.get("observed_citation_share")]
        cs.append(row)
        _style_numbers(cs, cluster_kinds)
    cs.freeze_panes = "A2"

    # Winnable sheet.
    winnable = view.get("winnable")
    if winnable:
        ws2 = wb.create_sheet("Winnable")
        _header_row(ws2, ["Intent", "Winnable low", "Winnable high"], brand)
        band_kinds: list[tuple[str, int] | None] = [None, _COUNT, _COUNT]
        for r in winnable["by_intent"]:
            ws2.append(
                [
                    _safe_cell(r["intent"].replace("_", " ")),
                    r["winnable_band"][0],
                    r["winnable_band"][1],
                ]
            )
            _style_numbers(ws2, band_kinds)
        ws2.append([])
        band = winnable.get("portfolio_winnable_band") or [None, None]
        ws2.append(["Portfolio", band[0], band[1]])
        _style_numbers(ws2, band_kinds)

    # Gaps sheet.
    gaps = view.get("gaps") or []
    if gaps:
        gs = wb.create_sheet("Gaps")
        _header_row(gs, ["Entity", "Demand", "Suggested cluster"], brand)
        for g in gaps:
            gs.append(
                [
                    _safe_cell(g["entity"]),
                    g.get("demand_volume"),
                    _safe_cell(g.get("suggested_cluster_head")),
                ]
            )
            _style_numbers(gs, [None, _COUNT, None])

    props = wb.properties
    props.creator = brand["author"]
    props.title = title_text
    props.created = timestamp
    props.modified = timestamp

    wb.save(str(out_path))
